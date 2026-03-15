"""
ניתוח סטטיסטי לאפקט גרנר — Garner Effect Statistical Analysis
================================================================
מריץ את כל הניתוחים הסטטיסטיים על נתוני הניסוי ושומר לקובץ Excel.

שימוש:
    python garner_statistical_analysis.py

פלט:
    Garner_Statistical_Analysis.xlsx
"""

import os, csv, io
import pandas as pd
import numpy as np
from scipy import stats
import pingouin as pg
import warnings
warnings.filterwarnings('ignore')

SUBJECTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Subjects')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Garner_Statistical_Analysis.xlsx')


def load_all_trials():
    """Load raw trial data from all subject CSV files."""
    all_trials = []
    for fname in sorted(os.listdir(SUBJECTS_DIR)):
        if not fname.startswith('Garner_Full_') or not fname.endswith('.csv'):
            continue
        with open(os.path.join(SUBJECTS_DIR, fname), 'r', encoding='utf-8-sig') as f:
            content = f.read()
        raw = content.split('=== RAW DATA ===')[1].strip()
        for row in csv.DictReader(io.StringIO(raw)):
            row['RT_ms'] = int(row['RT_ms'])
            row['Subject'] = int(row['Subject'])
            all_trials.append(row)
    return pd.DataFrame(all_trials)


def compute_aggregates(df):
    """Compute per-subject mean RT by condition (correct trials only)."""
    df_correct = df[df['Correction'] == 'correct'].copy()
    df_correct['RT_ms'] = df_correct['RT_ms'].astype(float)
    agg = df_correct.groupby(['Subject', 'Group', 'Stage', 'Congruency'])['RT_ms'].mean().reset_index()
    agg.rename(columns={'RT_ms': 'MeanRT'}, inplace=True)
    return agg


def analysis_rm_anova_stage_congruency(df_bf):
    """Two-way RM ANOVA: Stage × Congruency (all subjects)."""
    aov = pg.rm_anova(data=df_bf, dv='MeanRT', within=['Stage', 'Congruency'], subject='Subject')
    result = aov[['Source', 'ddof1', 'ddof2', 'F', 'p_unc', 'ng2', 'eps']].copy()
    result.columns = ['Source', 'df1', 'df2', 'F', 'p', 'η²g', 'ε']
    result['Significant'] = result['p'].apply(lambda x: '***' if x < .001 else ('**' if x < .01 else ('*' if x < .05 else 'n.s.')))
    return result


def paired_t_test(vals_a, vals_b, label_a, label_b, effect_name):
    """Run a paired-samples t-test and return a summary row."""
    diff = vals_b - vals_a
    t_val, p_val = stats.ttest_rel(vals_b, vals_a)
    d = diff.mean() / diff.std() if diff.std() > 0 else 0
    return {
        'Effect': effect_name,
        f'{label_a} M': round(vals_a.mean()),
        f'{label_a} SD': round(vals_a.std()),
        f'{label_b} M': round(vals_b.mean()),
        f'{label_b} SD': round(vals_b.std()),
        'Difference': round(diff.mean()),
        'df': len(vals_a) - 1,
        't': round(t_val, 3),
        'p': round(p_val, 6),
        "Cohen's d": round(d, 3),
        'Significant': '***' if p_val < .001 else ('**' if p_val < .01 else ('*' if p_val < .05 else 'n.s.'))
    }


def independent_t_test(vals_a, vals_b, label_a, label_b, effect_name):
    """Run an independent-samples t-test and return a summary row."""
    t_val, p_val = stats.ttest_ind(vals_a, vals_b)
    n_a, n_b = len(vals_a), len(vals_b)
    pooled_sd = np.sqrt(((n_a - 1) * vals_a.std()**2 + (n_b - 1) * vals_b.std()**2) / (n_a + n_b - 2))
    d = (vals_a.mean() - vals_b.mean()) / pooled_sd if pooled_sd > 0 else 0
    return {
        'Effect': effect_name,
        f'{label_a} M': round(vals_a.mean()),
        f'{label_a} SD': round(vals_a.std()),
        f'{label_b} M': round(vals_b.mean()),
        f'{label_b} SD': round(vals_b.std()),
        'Difference': round(vals_a.mean() - vals_b.mean()),
        'df': n_a + n_b - 2,
        't': round(t_val, 3),
        'p': round(p_val, 6),
        "Cohen's d": round(d, 3),
        'Significant': '***' if p_val < .001 else ('**' if p_val < .01 else ('*' if p_val < .05 else 'n.s.'))
    }


def run_all_analyses():
    """Run all analyses and write to Excel."""
    df = load_all_trials()
    df['RT_ms'] = df['RT_ms'].astype(float)
    agg = compute_aggregates(df)
    df_bf = agg[agg['Stage'].isin(['Baseline', 'Filtering'])].copy()

    n_total = df['Subject'].nunique()
    n_pitch = df[df['Group'] == 'PITCH']['Subject'].nunique()
    n_odor = df[df['Group'] == 'ODOR']['Subject'].nunique()

    sheets = {}

    # --- Sheet 1: Sample Info ---
    info = pd.DataFrame([{
        'Total N': n_total, 'PITCH N': n_pitch, 'ODOR N': n_odor,
        'Analysis': 'Garner Paradigm — Baseline vs Filtering, Congruent vs Incongruent',
        'Correct trials only': 'Yes',
        'DV': 'Mean RT (ms)'
    }])
    sheets['מידע כללי'] = info

    # --- Sheet 2: RM ANOVA Stage × Congruency ---
    anova_result = analysis_rm_anova_stage_congruency(df_bf)
    sheets['ANOVA Stage×Congruency'] = anova_result

    # --- Sheet 3: Main effects (paired t-tests) ---
    main_effects = []

    # Stage main effect (Garner Interference)
    stage_means = df_bf.groupby(['Subject', 'Stage'])['MeanRT'].mean().reset_index()
    bl = stage_means[stage_means['Stage'] == 'Baseline']['MeanRT'].values
    fl = stage_means[stage_means['Stage'] == 'Filtering']['MeanRT'].values
    main_effects.append(paired_t_test(bl, fl, 'Baseline', 'Filtering', 'Garner Interference (Stage)'))

    # Congruency main effect
    cong_means = df_bf.groupby(['Subject', 'Congruency'])['MeanRT'].mean().reset_index()
    cg = cong_means[cong_means['Congruency'] == 'congruent']['MeanRT'].values
    ig = cong_means[cong_means['Congruency'] == 'incongruent']['MeanRT'].values
    main_effects.append(paired_t_test(cg, ig, 'Congruent', 'Incongruent', 'Congruency Effect'))

    # Group main effect (independent)
    grp_means = df_bf.groupby(['Subject', 'Group'])['MeanRT'].mean().reset_index()
    pitch_rt = grp_means[grp_means['Group'] == 'PITCH']['MeanRT'].values
    odor_rt = grp_means[grp_means['Group'] == 'ODOR']['MeanRT'].values
    main_effects.append(independent_t_test(pitch_rt, odor_rt, 'PITCH', 'ODOR', 'Group (PITCH vs ODOR)'))

    sheets['אפקטים ראשיים'] = pd.DataFrame(main_effects)

    # --- Sheet 4: Garner by group ---
    garner_by_group = []
    for grp in ['PITCH', 'ODOR']:
        sm = df_bf[df_bf['Group'] == grp].groupby(['Subject', 'Stage'])['MeanRT'].mean().reset_index()
        bl_g = sm[sm['Stage'] == 'Baseline']['MeanRT'].values
        fl_g = sm[sm['Stage'] == 'Filtering']['MeanRT'].values
        row = paired_t_test(bl_g, fl_g, 'Baseline', 'Filtering', f'Garner Interference — {grp}')
        row['Group'] = grp
        garner_by_group.append(row)
    sheets['גרנר לפי קבוצה'] = pd.DataFrame(garner_by_group)

    # --- Sheet 5: Congruency by group ---
    cong_by_group = []
    for grp in ['PITCH', 'ODOR']:
        cm = df_bf[df_bf['Group'] == grp].groupby(['Subject', 'Congruency'])['MeanRT'].mean().reset_index()
        cg_g = cm[cm['Congruency'] == 'congruent']['MeanRT'].values
        ig_g = cm[cm['Congruency'] == 'incongruent']['MeanRT'].values
        row = paired_t_test(cg_g, ig_g, 'Congruent', 'Incongruent', f'Congruency — {grp}')
        row['Group'] = grp
        cong_by_group.append(row)
    sheets['התאמה לפי קבוצה'] = pd.DataFrame(cong_by_group)

    # --- Sheet 6: Interactions (between-group comparisons) ---
    interactions = []

    # Group × Stage
    gi_pitch = df_bf[df_bf['Group'] == 'PITCH'].groupby(['Subject', 'Stage'])['MeanRT'].mean().reset_index()
    gi_pitch_pv = gi_pitch.pivot(index='Subject', columns='Stage', values='MeanRT')
    gi_pitch_diff = (gi_pitch_pv['Filtering'] - gi_pitch_pv['Baseline']).values

    gi_odor = df_bf[df_bf['Group'] == 'ODOR'].groupby(['Subject', 'Stage'])['MeanRT'].mean().reset_index()
    gi_odor_pv = gi_odor.pivot(index='Subject', columns='Stage', values='MeanRT')
    gi_odor_diff = (gi_odor_pv['Filtering'] - gi_odor_pv['Baseline']).values

    interactions.append(independent_t_test(gi_pitch_diff, gi_odor_diff,
                                           'PITCH GI', 'ODOR GI', 'Group × Stage (GI difference)'))

    # Group × Congruency
    for grp in ['PITCH', 'ODOR']:
        cm2 = df_bf[df_bf['Group'] == grp].groupby(['Subject', 'Congruency'])['MeanRT'].mean().reset_index()
        pv = cm2.pivot(index='Subject', columns='Congruency', values='MeanRT')
        diff = (pv['incongruent'] - pv['congruent']).values
        if grp == 'PITCH':
            pitch_ce = diff
        else:
            odor_ce = diff

    interactions.append(independent_t_test(pitch_ce, odor_ce,
                                           'PITCH CE', 'ODOR CE', 'Group × Congruency (CE difference)'))

    sheets['אינטראקציות'] = pd.DataFrame(interactions)

    # --- Sheet 7: Congruency in Filtering only ---
    df_filt = df_bf[df_bf['Stage'] == 'Filtering']
    filt_cong = df_filt.groupby(['Subject', 'Congruency'])['MeanRT'].mean().reset_index()
    fc = filt_cong[filt_cong['Congruency'] == 'congruent']['MeanRT'].values
    fi = filt_cong[filt_cong['Congruency'] == 'incongruent']['MeanRT'].values
    filt_row = paired_t_test(fc, fi, 'Congruent', 'Incongruent', 'Congruency in Filtering stage')
    sheets['התאמה ב-Filtering'] = pd.DataFrame([filt_row])

    # --- Sheet 8: Descriptive statistics per condition ---
    desc = df_bf.groupby(['Group', 'Stage', 'Congruency'])['MeanRT'].agg(['mean', 'std', 'count']).reset_index()
    desc.columns = ['Group', 'Stage', 'Congruency', 'Mean RT', 'SD', 'N']
    desc['Mean RT'] = desc['Mean RT'].round(0).astype(int)
    desc['SD'] = desc['SD'].round(0).astype(int)
    sheets['סטטיסטיקה תיאורית'] = desc

    # --- Write to Excel ---
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        sig_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        ns_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Headers
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border
            # Data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    # Color significant column
                    if cell.value == '***' or cell.value == '**' or cell.value == '*':
                        cell.fill = sig_fill
                        cell.font = Font(bold=True, color='006100')
                    elif cell.value == 'n.s.':
                        cell.fill = ns_fill
                        cell.font = Font(color='9C0006')
            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(c.value or '')) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)

    print(f'Saved: {OUTPUT_FILE}')
    print(f'Sheets: {list(sheets.keys())}')


if __name__ == '__main__':
    run_all_analyses()
